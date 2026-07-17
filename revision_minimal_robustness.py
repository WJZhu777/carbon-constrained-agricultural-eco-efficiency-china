from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import KFold
from sklearn.preprocessing import StandardScaler
from xgboost import XGBRegressor


ROOT = Path(__file__).resolve().parent
DATA_PATH = ROOT / "data.xlsx"
TABLES_DIR = ROOT / "tables"

ID_COL = "ID"
YEAR_COL = "Year"
Y_COL = "efficiency"

FEATURES = ["TPAM", "EIA", "CS", "AFA", "PU", "ADY", "PFU", "NRP", "GAO", "CEA"]
CEA_COMPONENT_COLS = {"AFA", "PU", "ADY", "PFU", "EIA", "CS"}

PROVINCES_BY_ID = [
    "Beijing",
    "Tianjin",
    "Hebei",
    "Shanxi",
    "Inner Mongolia",
    "Liaoning",
    "Jilin",
    "Heilongjiang",
    "Shanghai",
    "Jiangsu",
    "Zhejiang",
    "Anhui",
    "Fujian",
    "Jiangxi",
    "Shandong",
    "Henan",
    "Hubei",
    "Hunan",
    "Guangdong",
    "Guangxi",
    "Hainan",
    "Chongqing",
    "Sichuan",
    "Guizhou",
    "Yunnan",
    "Shaanxi",
    "Gansu",
    "Qinghai",
    "Ningxia",
    "Xinjiang",
]

MONOTONE_BY_FEATURE = {
    "TPAM": -1,
    "EIA": -1,
    "CS": -1,
    "AFA": -1,
    "PU": -1,
    "ADY": -1,
    "PFU": -1,
    "NRP": -1,
    "GAO": 1,
    "CEA": -1,
}

def load_modeling_data() -> pd.DataFrame:
    df = pd.read_excel(DATA_PATH, sheet_name="Sheet1")
    need = [ID_COL, YEAR_COL, Y_COL] + FEATURES
    missing = [c for c in need if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns in data.xlsx: {missing}")
    df = df[need].copy()
    for c in need:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["Province"] = df[ID_COL].astype(int).map(
        lambda x: PROVINCES_BY_ID[x - 1] if 1 <= x <= len(PROVINCES_BY_ID) else f"ID_{x}"
    )
    df["row_key"] = np.arange(len(df), dtype=int)
    return df.dropna(subset=need).sort_values([YEAR_COL, ID_COL]).reset_index(drop=True)


def metrics_row(y_true, y_pred) -> dict[str, float]:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    mse = mean_squared_error(y_true, y_pred)
    return {
        "MSE": float(mse),
        "RMSE": float(np.sqrt(mse)),
        "MAE": float(mean_absolute_error(y_true, y_pred)),
        "R2": float(r2_score(y_true, y_pred)),
    }


def make_xgb(features: list[str], random_state: int = 42) -> XGBRegressor:
    constraints = tuple(MONOTONE_BY_FEATURE[f] for f in features)
    return XGBRegressor(
        n_estimators=500,
        max_depth=3,
        learning_rate=0.03,
        subsample=0.90,
        colsample_bytree=0.90,
        objective="reg:squarederror",
        reg_lambda=1.0,
        random_state=random_state,
        n_jobs=-1,
        monotone_constraints=constraints,
    )


def fit_predict_xgb(df: pd.DataFrame, features: list[str], train_idx, test_idx, seed: int) -> np.ndarray:
    x_train = df.loc[train_idx, features].to_numpy(dtype=float)
    y_train = df.loc[train_idx, Y_COL].to_numpy(dtype=float)
    x_test = df.loc[test_idx, features].to_numpy(dtype=float)
    scaler = StandardScaler()
    x_train_s = scaler.fit_transform(x_train)
    x_test_s = scaler.transform(x_test)
    model = make_xgb(features, random_state=seed)
    model.fit(x_train_s, y_train)
    return model.predict(x_test_s)


def format_output_sheets(writer: pd.ExcelWriter) -> None:
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        for column_index, cells in enumerate(worksheet.iter_cols(), start=1):
            max_length = max((len(str(cell.value)) for cell in cells if cell.value is not None), default=0)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 55)


def run_surrogate_predictor_ablation(df: pd.DataFrame) -> Path:
    out_path = TABLES_DIR / "Surrogate_predictor_ablation_revision.xlsx"
    feature_sets = {
        "full_predictors": FEATURES,
        "without_CEA": [c for c in FEATURES if c != "CEA"],
        "without_CEA_and_CEA_components": [
            c for c in FEATURES if c != "CEA" and c not in CEA_COMPONENT_COLS
        ],
    }

    rows = []
    kf = KFold(n_splits=5, shuffle=True, random_state=42)
    for feature_set, feats in feature_sets.items():
        for fold, (tr_pos, te_pos) in enumerate(kf.split(df), start=1):
            pred = fit_predict_xgb(df, feats, df.index[tr_pos], df.index[te_pos], seed=4200 + fold)
            m = metrics_row(df.loc[df.index[te_pos], Y_COL], pred)
            rows.append(
                {
                    "scheme": "random_5fold",
                    "split_id": fold,
                    "feature_set": feature_set,
                    "n_features": len(feats),
                    "train_n": int(len(tr_pos)),
                    "test_n": int(len(te_pos)),
                    "train_year_max": np.nan,
                    "test_years": "mixed",
                    **m,
                }
            )

    years = sorted(int(y) for y in df[YEAR_COL].unique())
    min_year, max_year = min(years), max(years)
    split_id = 0
    for train_end in range(min_year + 9, max_year - 3):
        val_year = train_end + 1
        test_years = [train_end + 2, train_end + 3, train_end + 4]
        if not all(y in years for y in test_years):
            continue
        tr_idx = df.index[df[YEAR_COL] <= train_end]
        te_idx = df.index[df[YEAR_COL].isin(test_years)]
        if len(tr_idx) == 0 or len(te_idx) == 0:
            continue
        split_id += 1
        for feature_set, feats in feature_sets.items():
            pred = fit_predict_xgb(df, feats, tr_idx, te_idx, seed=5200 + split_id)
            m = metrics_row(df.loc[te_idx, Y_COL], pred)
            rows.append(
                {
                    "scheme": "time_forward_rolling",
                    "split_id": split_id,
                    "feature_set": feature_set,
                    "n_features": len(feats),
                    "train_n": int(len(tr_idx)),
                    "test_n": int(len(te_idx)),
                    "train_year_max": int(train_end),
                    "validation_year_gap": int(val_year),
                    "test_years": ",".join(str(y) for y in test_years),
                    **m,
                }
            )

    detail = pd.DataFrame(rows)
    summary = (
        detail.groupby(["scheme", "feature_set"], as_index=False)
        .agg(
            n_splits=("split_id", "count"),
            n_features=("n_features", "first"),
            MSE_mean=("MSE", "mean"),
            MSE_std=("MSE", "std"),
            RMSE_mean=("RMSE", "mean"),
            MAE_mean=("MAE", "mean"),
            R2_mean=("R2", "mean"),
            R2_std=("R2", "std"),
        )
        .sort_values(["scheme", "R2_mean"], ascending=[True, False])
    )
    readme = pd.DataFrame(
        {
            "Notes": [
                "Reviewer-driven surrogate ablation for circularity concerns.",
                "Model is monotone-constrained XGBoost using the same predictor signs as the main pipeline.",
                "without_CEA removes only the aggregate carbon-emission variable.",
                "without_CEA_and_CEA_components removes CEA plus AFA, PU, ADY, PFU, EIA, and CS.",
                "The analysis is diagnostic: it evaluates label approximation under reduced predictor sets, not causal mechanisms.",
            ]
        }
    )
    fs = pd.DataFrame(
        [{"feature_set": k, "features": ", ".join(v), "n_features": len(v)} for k, v in feature_sets.items()]
    )
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="split_detail", index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)
        fs.to_excel(writer, sheet_name="feature_sets", index=False)
        readme.to_excel(writer, sheet_name="ReadMe", index=False)
        format_output_sheets(writer)
    return out_path


def main() -> None:
    TABLES_DIR.mkdir(exist_ok=True)
    df = load_modeling_data()
    print(f"[revision robustness] loaded {len(df)} rows from {DATA_PATH}")
    path = run_surrogate_predictor_ablation(df)
    print(f"[revision robustness] saved {path}")


if __name__ == "__main__":
    main()
